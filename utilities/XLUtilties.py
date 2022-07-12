import openpyxl

def getRowcount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_row)

def getColumncount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_column)


def readdata(file, sheetname, rownumber, columnnumber):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row = rownumber, column = columnnumber).value


def writtendata(file, sheetname, rownumber, columnnumber, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownumber, column=columnnumber).value = data
    workbook.save(file)